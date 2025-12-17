"""
バックテスト結果の視覚化ツール
"""

import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.patches import Rectangle
from backtest import BacktestEngine, select_csv_file
from openpyxl.drawing.image import Image as XLImage
from openpyxl import load_workbook
import os
import re

plt.rcParams['font.sans-serif'] = ['MS Gothic', 'Yu Gothic', 'Meiryo']
plt.rcParams['axes.unicode_minus'] = False


class BacktestVisualizer:
    
    def __init__(self, csv_path):
        self.csv_path = csv_path
        self.bt = BacktestEngine(csv_path)
        
    def run_backtest(self):
        print("バックテスト実行中...")
        self.bt.run()
        print("完了！")
        
    def export_to_excel_with_chart(self, excel_path='backtest_result.xlsx'):
        if len(self.bt.trades) == 0:
            print("取引データがありません")
            return
        
        temp_chart_path = 'temp_chart.png'
        self._create_chart_for_excel(temp_chart_path)
        
        trades_df = pd.DataFrame(self.bt.trades)
        
        trades_df['entry_time'] = pd.to_datetime(trades_df['entry_time']).dt.tz_localize(None).dt.strftime('%Y-%m-%d %H:%M:%S')
        trades_df['exit_time'] = pd.to_datetime(trades_df['exit_time']).dt.tz_localize(None).dt.strftime('%Y-%m-%d %H:%M:%S')
        
        trades_df['累積損益_pips'] = trades_df['pips'].cumsum()
        
        trades_df_jp = pd.DataFrame({
            'エントリー日時': trades_df['entry_time'],
            '決済日時': trades_df['exit_time'],
            '方向': trades_df['direction'],
            'エントリー価格': trades_df['entry_price'],
            '決済価格': trades_df['exit_price'],
            '獲得pips': trades_df['pips'],
            '累積損益': trades_df['累積損益_pips']
        })
        
        metrics = self.bt.calculate_metrics()
        metrics_df = pd.DataFrame([metrics])
        metrics_df.columns = ['総損益_pips', '取引回数', '勝率_%', '最大DD_pips', 'プロフィットファクター']
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            metrics_df.to_excel(writer, sheet_name='サマリー', index=False)
            trades_df_jp.to_excel(writer, sheet_name='取引一覧', index=False)
        
        wb = load_workbook(excel_path)
        ws_chart = wb.create_sheet('チャート')
        
        img = XLImage(temp_chart_path)
        img.width = 2400
        img.height = 960
        ws_chart.add_image(img, 'A1')
        
        wb.save(excel_path)
        
        if os.path.exists(temp_chart_path):
            os.remove(temp_chart_path)
        
        print(f"✅ Excelファイルを作成しました（チャート埋め込み済み）: {excel_path}")
        
    def _create_chart_for_excel(self, output_path):
        df = self.bt.load_data()
        df = self.bt.calculate_sma(df)
        df = self.bt.calculate_heikin_ashi(df)
        df = self.bt.calculate_ha_color(df)
        df = self.bt.calculate_ha_body(df)
        
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(32, 12), 
                                        gridspec_kw={'height_ratios': [3, 1]})
        
        self._plot_candlesticks(ax1, df)
        ax1.plot(df['time'], df['sma75'], label='75SMA', color='blue', linewidth=2)
        self._plot_heikin_ashi_background(ax1, df)
        self._plot_trade_points(ax1, df)
        
        ax1.set_title('バックテスト結果 - USD/JPY 1時間足', fontsize=16, fontweight='bold')
        ax1.set_ylabel('価格', fontsize=12)
        ax1.legend(loc='upper left')
        ax1.grid(True, alpha=0.3)
        ax1.xaxis.set_major_formatter(mdates.DateFormatter('%m/%d'))
        
        if len(self.bt.trades) > 0:
            trades_df = pd.DataFrame(self.bt.trades)
            trades_df['cumulative_pips'] = trades_df['pips'].cumsum()
            exit_times = pd.to_datetime(trades_df['exit_time']).dt.tz_localize(None)
            
            ax2.plot(exit_times, trades_df['cumulative_pips'], 
                     color='green', linewidth=2, marker='o')
            ax2.axhline(y=0, color='red', linestyle='--', alpha=0.5)
            ax2.fill_between(exit_times, 0, trades_df['cumulative_pips'], 
                            alpha=0.3, color='green')
            
            ax2.set_title('累積損益', fontsize=14)
            ax2.set_xlabel('日時', fontsize=12)
            ax2.set_ylabel('損益 (pips)', fontsize=12)
            ax2.grid(True, alpha=0.3)
            ax2.xaxis.set_major_formatter(mdates.DateFormatter('%m/%d'))
        
        plt.tight_layout()
        plt.savefig(output_path, dpi=150, bbox_inches='tight')
        plt.close()
    
    def _plot_candlesticks(self, ax, df):
        width = 1.8 / 24
        
        for idx, row in df.iterrows():
            color = 'red' if row['close'] >= row['open'] else 'blue'
            height = abs(row['close'] - row['open'])
            bottom = min(row['open'], row['close'])
            rect = Rectangle((mdates.date2num(row['time']) - width/2, bottom),
                           width, height, facecolor=color, edgecolor='black', 
                           alpha=0.7, linewidth=0.5)
            ax.add_patch(rect)
            ax.plot([row['time'], row['time']], [row['low'], row['high']], 
                   color='black', linewidth=0.5)
    
    def _plot_heikin_ashi_background(self, ax, df):
        for idx in range(len(df)):
            color = 'lightcoral' if df['ha_color'].iloc[idx] == 'blue' else 'lightblue'
            
            if idx < len(df) - 1:
                time_start = df['time'].iloc[idx]
                time_end = df['time'].iloc[idx + 1]
            else:
                time_start = df['time'].iloc[idx]
                time_end = df['time'].iloc[idx] + pd.Timedelta(hours=1)
            
            ax.axvspan(time_start, time_end, alpha=0.15, color=color)
    
    def _plot_trade_points(self, ax, df):
        for trade in self.bt.trades:
            entry_time = pd.to_datetime(trade['entry_time']).tz_localize(None)
            exit_time = pd.to_datetime(trade['exit_time']).tz_localize(None)
            entry_price = trade['entry_price']
            exit_price = trade['exit_price']
            direction = trade['direction']
            pips = trade['pips']
            
            marker = '^' if direction == 'long' else 'v'
            color = 'green' if direction == 'long' else 'red'
            
            ax.scatter(entry_time, entry_price, marker=marker, 
                      s=200, color=color, edgecolor='black', 
                      linewidth=1.5, zorder=5, alpha=0.8)
            
            profit_color = 'green' if pips > 0 else 'red'
            exit_marker = 'o' if pips > 0 else 'x'
            
            if pips > 0:
                ax.scatter(exit_time, exit_price, marker=exit_marker, 
                          s=300, facecolors='none', edgecolors=profit_color, 
                          linewidth=2.5, zorder=5)
            else:
                ax.scatter(exit_time, exit_price, marker=exit_marker, 
                          s=200, color=profit_color, 
                          linewidth=3, zorder=5)
            
            ax.plot([entry_time, exit_time], [entry_price, exit_price], 
                   color=profit_color, linestyle='--', alpha=0.5, linewidth=1)


def generate_output_filename(csv_filename):
    match = re.search(r'(\d{4})-(\d{2})', csv_filename)
    
    if match:
        year = match.group(1)
        month = match.group(2)
        return f'backtest_{year}-{month}.xlsx'
    else:
        return 'backtest_result.xlsx'


if __name__ == "__main__":
    csv_path = select_csv_file()
    
    if csv_path:
        filename = os.path.basename(csv_path)
        output_name = generate_output_filename(filename)
        
        print(f"\n選択されたファイル: {filename}")
        print(f"出力ファイル名: {output_name}\n")
        
        viz = BacktestVisualizer(csv_path)
        viz.run_backtest()
        viz.bt.print_results()
        viz.export_to_excel_with_chart(output_name)
        
        print(f"\n✅ すべての処理が完了しました！")
        print(f"📊 {output_name}")